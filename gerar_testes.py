"""
gerar_testes.py
─────────────────────────────────────────────────────────────
Script para gerar (ou regenerar) a aba "Testes de Aceitação"
em planilhas Excel já criadas pelo backlog_agent.py.

Funciona com qualquer Excel do backlog_agent — projetos novos ou mudanças.
Não altera nenhuma outra aba do arquivo.

Uso:
    python gerar_testes.py

Dependências:
    pip install anthropic openpyxl python-dotenv
"""

import os, json, re, datetime
from pathlib import Path
from dotenv import load_dotenv
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()
CLIENT = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
MODEL  = "claude-haiku-4-5-20251001"

# ── Utilitários de terminal ────────────────────────────────
AZUL    = "\033[94m"
VERDE   = "\033[92m"
AMARELO = "\033[93m"
CINZA   = "\033[90m"
RESET   = "\033[0m"
BOLD    = "\033[1m"

def titulo(t):   print(f"\n{BOLD}{AZUL}{'═'*60}\n  {t}\n{'═'*60}{RESET}")
def secao(t):    print(f"\n{BOLD}{VERDE}── {t}{RESET}")
def info(t):     print(f"{CINZA}   {t}{RESET}")
def pergunta(t): return input(f"\n{AMARELO}▶  {t}{RESET} ").strip()
def ok(t):       print(f"{VERDE}   ✓ {t}{RESET}")
def aviso(t):    print(f"{AMARELO}   ⚠ {t}{RESET}")

# ── Estilos Excel ──────────────────────────────────────────
def fill_xl(h): return PatternFill("solid", start_color=h, fgColor=h)
def fnt_xl(bold=False, size=10, color="222222"):
    return Font(bold=bold, size=size, color=color, name="Arial")
def aln_xl(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd_xl():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

MODULO_CORES = [
    "DEEAF1","EBF3FB","E8F5E3","F3E8FF","FFF4EC","E8EAF6",
    "DDEEFF","FFE8D6","E8FFE8","F0E8FF","FFEEDD","E0F0FF",
]

# ── Prompts ────────────────────────────────────────────────

PROMPT_TESTES_SIMPLES = """\
Você é um analista de qualidade (QA) especializado em testes de aceitação para software.

Sua tarefa é gerar os testes de aceitação para a User Story abaixo.

Regras obrigatórias:
- Gere APENAS os testes realmente necessários para validar a entrega — sem redundância, sem testes genéricos.
- Cada teste deve validar um comportamento específico e observável desta User Story, não de sistemas em geral.
- Se a US for simples (1 comportamento), gere 1 ou 2 testes. Se for complexa, gere até 5.
- Evite frases vagas como "verificar se o sistema funciona" ou "confirmar que não há erros".
- Use nomes reais de telas, campos, botões ou regras quando puder inferir pelo título e épico.
- Formato obrigatório: "Dado [contexto específico], quando [ação exata], então [resultado mensurável]."
- Os testes devem ser executáveis manualmente pelo analista de requisitos, sem acesso ao código.

User Story ID: {us_id}
Título: {titulo}
Épico: {epico}

Retorne APENAS um JSON válido, sem markdown, sem backticks:
{{
  "testes": [
    "Dado ..., quando ..., então ..."
  ]
}}
"""

PROMPT_TESTES_COMPLETO = """\
Você é um analista de qualidade (QA) especializado em testes de aceitação para software.

Sua tarefa é gerar os testes de aceitação para a User Story abaixo, com base nos critérios fornecidos.

Regras obrigatórias:
- Gere APENAS os testes realmente necessários — elimine redundâncias, não force quantidade.
- Cada critério de aceitação pode virar 1 teste direto; só divida em 2 se cobrir situações distintas (ex: sucesso e falha).
- Seja específico: use os nomes de campos, telas, regras e valores que aparecem nos critérios.
- Evite testes genéricos. Cada teste deve ser inconfundível com outro US do projeto.
- Formato obrigatório: "Dado [contexto específico], quando [ação exata], então [resultado mensurável]."
- Os testes devem ser executáveis manualmente pelo analista de requisitos, sem acesso ao código.

User Story ID: {us_id}
Título: {titulo}
História: {historia}
Critérios de Aceitação:
{criterios}

Retorne APENAS um JSON válido, sem markdown, sem backticks:
{{
  "testes": [
    "Dado ..., quando ..., então ..."
  ]
}}
"""

# ── Chamada à API ──────────────────────────────────────────

def chamar_claude(prompt: str) -> str:
    response = CLIENT.messages.create(
        model=MODEL,
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.content[0].text

def extrair_json(texto: str) -> dict:
    texto = texto.strip()
    texto = re.sub(r"```json\s*", "", texto)
    texto = re.sub(r"```\s*", "", texto)
    inicio = texto.find("{")
    fim    = texto.rfind("}") + 1
    if inicio == -1 or fim == 0:
        raise ValueError("Nenhum JSON encontrado na resposta.")
    return json.loads(texto[inicio:fim])

# ── Leitura do Excel ───────────────────────────────────────

def ler_backlog_do_excel(caminho: str) -> list:
    """
    Lê a aba 'Backlog' do Excel gerado pelo backlog_agent.
    Estrutura esperada (linha 2 = cabeçalhos, linha 3+ = dados):
      Col A: Sprint | Col B: Semana | Col C: ID | Col D: Título
      Col E: Épico  | Col F: RF    | ...
    Retorna lista de dicts com dados de cada US única.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True)
    if "Backlog" not in wb.sheetnames:
        wb.close()
        raise ValueError(
            "Aba 'Backlog' não encontrada.\n"
            "Certifique-se de usar um Excel gerado pelo backlog_agent.py."
        )

    ws = wb["Backlog"]
    us_list = []
    seen_ids = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        # Ignora linhas completamente vazias
        if not any(row):
            continue

        sprint_val = str(row[0]).strip() if row[0] else ""
        us_id      = str(row[2]).strip() if row[2] else ""
        titulo_val = str(row[3]).strip() if row[3] else ""
        epico_val  = str(row[4]).strip() if row[4] else ""

        if not us_id or us_id in seen_ids:
            continue

        seen_ids.add(us_id)
        us_list.append({
            "sprint":  sprint_val,
            "id":      us_id,
            "titulo":  titulo_val,
            "epico":   epico_val,
        })

    wb.close()
    return us_list

# ── Geração de testes via Claude ───────────────────────────

def gerar_testes_para_us(us: dict) -> list:
    """Gera testes de aceitação para uma US. Usa prompt simples (sem criterios)."""
    prompt = PROMPT_TESTES_SIMPLES.format(
        us_id=us["id"],
        titulo=us["titulo"],
        epico=us.get("epico", ""),
    )
    try:
        resposta = chamar_claude(prompt)
        dados    = extrair_json(resposta)
        testes   = dados.get("testes", [])
        if not testes:
            raise ValueError("Lista de testes vazia.")
        return testes
    except Exception as e:
        aviso(f"Erro ao gerar testes para {us['id']}: {e}")
        return [
            f"Verificar manualmente se a funcionalidade '{us['titulo']}' foi entregue conforme descrito.",
            f"Confirmar que não há erros ou comportamentos inesperados na funcionalidade '{us['titulo']}'.",
        ]

# ── Criação da aba no Excel ────────────────────────────────

RESULTADO_OPCOES = "⬜ Não executado,✅ Passou,❌ Falhou,🔄 Reteste necessário"

def criar_aba_testes(wb, us_testes: list, projeto_nome: str = "Projeto") -> None:
    """
    Cria (ou substitui) a aba 'Testes de Aceitação' no workbook.
    Cada linha = 1 teste de 1 US. Inclui dropdown de resultado e auto-filter.
    """
    # Remove aba se já existir (regeneração limpa)
    if "Testes de Aceitação" in wb.sheetnames:
        del wb["Testes de Aceitação"]

    ws = wb.create_sheet("Testes de Aceitação")
    ws.freeze_panes = "A3"

    # ── Título ─────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ct = ws["A1"]
    ct.value = (
        f"TESTES DE ACEITAÇÃO · {projeto_nome.upper()}  |  "
        f"Preencha após cada entrega do sprint  |  "
        f"Gerado em {datetime.date.today().strftime('%d/%m/%Y')}"
    )
    ct.font      = fnt_xl(bold=True, size=11, color="FFFFFF")
    ct.fill      = fill_xl("1F4E79")
    ct.alignment = aln_xl(h="center", v="center")
    ws.row_dimensions[1].height = 28

    # ── Cabeçalhos ─────────────────────────────────────────
    COLS = [
        "Sprint", "US ID", "Título da US", "# Teste",
        "Descrição do Teste de Aceitação",
        "Resultado", "Observação do Analista", "Data", "Testador",
    ]
    WIDS = [9, 9, 34, 8, 62, 24, 36, 13, 18]

    for i, (col, w) in enumerate(zip(COLS, WIDS), 1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font      = fnt_xl(bold=True, color="FFFFFF")
        cell.fill      = fill_xl("2E75B6")
        cell.alignment = aln_xl(h="center", v="center", wrap=True)
        cell.border    = brd_xl()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    # ── Dados ──────────────────────────────────────────────
    row_i       = 3
    sprint_cores = {}
    cor_idx     = 0

    for us_data in us_testes:
        sprint = us_data.get("sprint", "")
        if sprint not in sprint_cores:
            sprint_cores[sprint] = MODULO_CORES[cor_idx % len(MODULO_CORES)]
            cor_idx += 1
        bg = sprint_cores[sprint]

        for num_teste, descricao in enumerate(us_data.get("testes", []), 1):
            # Colunas fixas (A–E)
            fixos = [sprint, us_data["id"], us_data["titulo"], num_teste, descricao]
            for col_i, val in enumerate(fixos, 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.fill      = fill_xl(bg)
                cell.border    = brd_xl()
                cell.font      = fnt_xl(size=9, bold=(col_i in [1, 4]))
                cell.alignment = aln_xl(
                    h="center" if col_i in [1, 2, 4] else "left",
                    v="center", wrap=True
                )

            # Resultado (F=6) — preenchido pelo analista
            rc = ws.cell(row=row_i, column=6, value="⬜ Não executado")
            rc.font      = fnt_xl(size=9, color="444444")
            rc.fill      = fill_xl("F0F0F0")
            rc.alignment = aln_xl(h="center", v="center")
            rc.border    = brd_xl()

            # Observação (G=7)
            for col_i in [7, 8, 9]:
                cell = ws.cell(row=row_i, column=col_i, value="")
                cell.fill      = fill_xl("FAFAFA")
                cell.border    = brd_xl()
                cell.alignment = aln_xl(
                    h="center" if col_i in [8, 9] else "left",
                    v="center", wrap=True
                )

            ws.row_dimensions[row_i].height = 36
            row_i += 1

    # ── Dropdown de resultado ──────────────────────────────
    dv = DataValidation(
        type="list",
        formula1=f'"{RESULTADO_OPCOES}"',
        showDropDown=False,   # False = mostra a setinha
        allow_blank=True,
    )
    dv.sqref = f"F3:F{row_i - 1}"
    ws.add_data_validation(dv)

    # ── Auto-filtro ────────────────────────────────────────
    ws.auto_filter.ref = f"A2:I{row_i - 1}"

    # ── Legenda de cores (abaixo dos dados) ───────────────
    leg_row = row_i + 1
    ws.merge_cells(f"A{leg_row}:I{leg_row}")
    lc = ws.cell(row=leg_row, column=1,
                 value="LEGENDA DO RESULTADO:  ⬜ Não executado  |  ✅ Passou  |  ❌ Falhou  |  🔄 Reteste necessário")
    lc.font      = fnt_xl(bold=False, size=9, color="FFFFFF")
    lc.fill      = fill_xl("2E75B6")
    lc.alignment = aln_xl(h="center", v="center")
    ws.row_dimensions[leg_row].height = 20


# ── Main ───────────────────────────────────────────────────

def main():
    titulo("GERADOR DE TESTES DE ACEITAÇÃO")
    info("Lê um Excel gerado pelo backlog_agent.py e adiciona (ou atualiza)")
    info("a aba 'Testes de Aceitação' sem alterar as demais abas.")
    info(f"Modelo: {MODEL} via API Anthropic\n")

    # ── 1. Caminho do Excel ────────────────────────────────
    while True:
        caminho = pergunta("Caminho do arquivo Excel (.xlsx):").strip('"').strip("'")
        if not caminho:
            aviso("Informe o caminho do arquivo.")
            continue
        p = Path(caminho)
        if not p.exists():
            aviso(f"Arquivo não encontrado: {caminho}")
            continue
        if p.suffix.lower() != ".xlsx":
            aviso("Apenas arquivos .xlsx são suportados.")
            continue
        break

    # ── 2. Ler US do Backlog ───────────────────────────────
    secao("Lendo aba Backlog...")
    try:
        us_list = ler_backlog_do_excel(caminho)
    except Exception as e:
        aviso(f"Erro ao ler o arquivo: {e}")
        return

    if not us_list:
        aviso("Nenhuma User Story encontrada na aba Backlog.")
        return

    ok(f"{len(us_list)} User Stories encontradas:")
    for us in us_list:
        info(f"  {us['sprint']:5s} | {us['id']:8s} — {us['titulo'][:55]}")

    confirmar = pergunta("\nContinuar e gerar os testes de aceitação? (s/n):").lower()
    if confirmar != "s":
        info("Operação cancelada.")
        return

    # ── 3. Nome do projeto ─────────────────────────────────
    nome_projeto = pergunta("Nome do projeto (para o título da aba):").strip() or "Projeto"

    # ── 4. Gerar testes via Claude ─────────────────────────
    secao(f"Gerando testes via IA ({MODEL})...")
    info("Isso pode levar alguns segundos por User Story...\n")

    us_testes = []
    for i, us in enumerate(us_list, 1):
        label = f"[{i:02d}/{len(us_list):02d}] {us['id']} — {us['titulo'][:45]}"
        info(f"{label}...")
        testes = gerar_testes_para_us(us)
        us["testes"] = testes
        us_testes.append(us)
        ok(f"{us['id']} → {len(testes)} testes gerados")

    total_testes = sum(len(u["testes"]) for u in us_testes)

    # ── 5. Adicionar aba ao Excel ──────────────────────────
    secao("Atualizando o Excel...")
    try:
        wb = openpyxl.load_workbook(caminho)
        criar_aba_testes(wb, us_testes, nome_projeto)
        wb.save(caminho)
        wb.close()
    except Exception as e:
        aviso(f"Erro ao salvar o arquivo: {e}")
        return

    # ── 6. Resultado ───────────────────────────────────────
    titulo("CONCLUÍDO!")
    print(f"""
  {VERDE}Arquivo atualizado:{RESET}
    📊  {caminho}

  {CINZA}Resumo:{RESET}
    • {len(us_list)} User Stories processadas
    • {total_testes} testes de aceitação gerados
    • Aba adicionada: 'Testes de Aceitação'

  {CINZA}Próximos passos:{RESET}
    1. Abra o Excel → aba 'Testes de Aceitação'
    2. Compartilhe com o analista de requisitos
    3. Após cada entrega do dev, preencher:
         - Coluna F (Resultado): ✅ Passou / ❌ Falhou / 🔄 Reteste necessário
         - Coluna G (Observação): detalhes do teste
         - Coluna H (Data): data da execução
         - Coluna I (Testador): nome do analista
    4. Use o filtro na coluna F para filtrar por status
""")


if __name__ == "__main__":
    main()
